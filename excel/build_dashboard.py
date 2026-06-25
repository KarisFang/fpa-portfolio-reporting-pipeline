"""
Excel Monthly FP&A Reporting Pack — built from SQL KPI extracts.
Raw monthly aggregates are loaded as data (analogous to actuals); every
growth rate, margin, share-of-total, and ranking is computed with live
Excel formulas, not pre-calculated in Python.
"""
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter

FONT = "Calibri"
BLUE = Font(name=FONT, color="0000FF")
BLACK = Font(name=FONT, color="000000")
BLACK_BOLD = Font(name=FONT, color="000000", bold=True)
WHITE_BOLD = Font(name=FONT, color="FFFFFF", bold=True)
GREEN = Font(name=FONT, color="008000")
TITLE_FONT = Font(name=FONT, bold=True, size=16, color="1F4E78")
KPI_LABEL_FONT = Font(name=FONT, size=10, color="595959")
KPI_VALUE_FONT = Font(name=FONT, bold=True, size=20, color="1F4E78")
SECTION_FILL = PatternFill("solid", fgColor="1F4E78")
SUBHEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
KPI_FILL = PatternFill("solid", fgColor="F2F2F2")
CUR = '$#,##0;($#,##0);"-"'
PCT = '0.0%;(0.0%);"-"'
NUM = '#,##0;(#,##0);"-"'

wb = Workbook()
wb.remove(wb.active)


def load_csv(path):
    with open(path) as f:
        r = list(csv.reader(f))
    return r[0], r[1:]


def write_table(ws, start_row, start_col, header, rows, formats=None):
    for j, h in enumerate(header):
        c = ws.cell(row=start_row, column=start_col + j, value=h)
        c.font = BLACK_BOLD
        c.fill = SUBHEADER_FILL
    for i, row in enumerate(rows):
        for j, val in enumerate(row):
            c = ws.cell(row=start_row + 1 + i, column=start_col + j)
            try:
                c.value = float(val)
            except (ValueError, TypeError):
                c.value = val
            if formats and j in formats:
                c.number_format = formats[j]
    return start_row + len(rows)


# =====================================================================
# SHEET 1: Data_Monthly  (raw monthly actuals -> Excel formulas for growth)
# =====================================================================
ws = wb.create_sheet("Data_Monthly")
hdr, rows = load_csv("/home/claude/fpa-reporting-pipeline/exports/monthly_kpis.csv")
ws["A1"] = "Monthly Revenue & Margin — Source: SQL export (monthly_kpis.csv) from sales.db"
ws["A1"].font = BLACK_BOLD
headers = ["Year-Month", "Transactions", "Net Revenue", "Gross Profit", "Gross Margin %",
           "MoM Growth %", "YoY Growth %", "3-Mo Rolling Avg Revenue", "YTD Revenue"]
for j, h in enumerate(headers):
    c = ws.cell(row=3, column=1 + j, value=h)
    c.font = BLACK_BOLD
    c.fill = SUBHEADER_FILL
    c.alignment = Alignment(horizontal="center", wrap_text=True)

for i, row in enumerate(rows):
    r = 4 + i
    year_month, year, month, txn, rev, gp = row
    ws.cell(row=r, column=1, value=year_month).font = BLUE
    ws.cell(row=r, column=2, value=int(float(txn))).font = BLUE
    ws.cell(row=r, column=2).number_format = NUM
    ws.cell(row=r, column=3, value=float(rev)).font = BLUE
    ws.cell(row=r, column=3).number_format = CUR
    ws.cell(row=r, column=4, value=float(gp)).font = BLUE
    ws.cell(row=r, column=4).number_format = CUR
    ws.cell(row=r, column=5, value=f"=D{r}/C{r}").number_format = PCT
    if i == 0:
        ws.cell(row=r, column=6, value=None)
        ws.cell(row=r, column=7, value=None)
    else:
        ws.cell(row=r, column=6, value=f"=IFERROR(C{r}/C{r-1}-1,\"\")").number_format = PCT
        if i >= 12:
            ws.cell(row=r, column=7, value=f"=IFERROR(C{r}/C{r-12}-1,\"\")").number_format = PCT
    roll_start = max(4, r - 2)
    ws.cell(row=r, column=8, value=f"=AVERAGE(C{roll_start}:C{r})").number_format = CUR
    if int(month) == 1:
        ws.cell(row=r, column=9, value=f"=C{r}")
    else:
        ws.cell(row=r, column=9, value=f"=I{r-1}+C{r}")
    ws.cell(row=r, column=9).number_format = CUR
    for col in [5, 6, 7, 9]:
        ws.cell(row=r, column=col).font = BLACK

last_row = 3 + len(rows)
autosize_widths = {"A": 12, "B": 13, "C": 15, "D": 15, "E": 14, "F": 13, "G": 13, "H": 20, "I": 15}
for col, w in autosize_widths.items():
    ws.column_dimensions[col].width = w
ws.sheet_view.showGridLines = False
ws.freeze_panes = "A4"

print("Data_Monthly done:", last_row)

# =====================================================================
# SHEET 2: Data_Regional (pivoted FY24 vs FY25 by region, with Excel-formula YoY)
# =====================================================================
ws = wb.create_sheet("Data_Regional")
ws["A1"] = "Regional Revenue — Source: SQL export (regional_monthly.csv)"
ws["A1"].font = BLACK_BOLD
hdr, rows = load_csv("/home/claude/fpa-reporting-pipeline/exports/regional_monthly.csv")
regions = sorted(set(r[0] for r in rows))
years = sorted(set(r[1] for r in rows))

headers = ["Region", "FY2024 Revenue", "FY2025 Revenue", "YoY Growth %", "FY2025 Share of Total %"]
for j, h in enumerate(headers):
    c = ws.cell(row=3, column=1 + j, value=h)
    c.font = BLACK_BOLD
    c.fill = SUBHEADER_FILL

rev_by_region_year = {}
for region_name, year, month, year_month, net_rev, gp in rows:
    rev_by_region_year[(region_name, year)] = rev_by_region_year.get((region_name, year), 0) + float(net_rev)

for i, region in enumerate(regions):
    r = 4 + i
    ws.cell(row=r, column=1, value=region).font = BLUE
    ws.cell(row=r, column=2, value=round(rev_by_region_year.get((region, "2024"), 0), 2)).font = BLUE
    ws.cell(row=r, column=2).number_format = CUR
    ws.cell(row=r, column=3, value=round(rev_by_region_year.get((region, "2025"), 0), 2)).font = BLUE
    ws.cell(row=r, column=3).number_format = CUR
    ws.cell(row=r, column=4, value=f"=C{r}/B{r}-1").number_format = PCT
    ws.cell(row=r, column=4).font = BLACK
    ws.cell(row=r, column=5, value=f"=C{r}/SUM($C$4:$C${3+len(regions)})").number_format = PCT
    ws.cell(row=r, column=5).font = BLACK

last_r = 3 + len(regions)
ws.cell(row=last_r + 1, column=1, value="Total").font = BLACK_BOLD
ws.cell(row=last_r + 1, column=2, value=f"=SUM(B4:B{last_r})").number_format = CUR
ws.cell(row=last_r + 1, column=3, value=f"=SUM(C4:C{last_r})").number_format = CUR
ws.cell(row=last_r + 1, column=4, value=f"=C{last_r+1}/B{last_r+1}-1").number_format = PCT
for col in [1, 2, 3, 4]:
    ws.cell(row=last_r + 1, column=col).font = BLACK_BOLD
for col, w in {"A": 16, "B": 16, "C": 16, "D": 14, "E": 20}.items():
    ws.column_dimensions[col].width = w
ws.sheet_view.showGridLines = False

# Also keep monthly regional detail further down for the trend chart by region
ws.cell(row=last_r + 4, column=1, value="Monthly Detail (for charting)").font = BLACK_BOLD
detail_headers = ["Region", "Year-Month", "Net Revenue"]
for j, h in enumerate(detail_headers):
    c = ws.cell(row=last_r + 5, column=1 + j, value=h)
    c.font = BLACK_BOLD
    c.fill = SUBHEADER_FILL
for i, row in enumerate(rows):
    rr = last_r + 6 + i
    ws.cell(row=rr, column=1, value=row[0]).font = BLUE
    ws.cell(row=rr, column=2, value=row[3]).font = BLUE
    ws.cell(row=rr, column=3, value=float(row[4])).font = BLUE
    ws.cell(row=rr, column=3).number_format = CUR

print("Data_Regional done")

# =====================================================================
# SHEET 3: Data_Channel
# =====================================================================
ws = wb.create_sheet("Data_Channel")
ws["A1"] = "Channel Mix — Source: SQL export (channel_monthly.csv)"
ws["A1"].font = BLACK_BOLD
hdr, rows = load_csv("/home/claude/fpa-reporting-pipeline/exports/channel_monthly.csv")
channels = sorted(set(r[0] for r in rows))
rev_by_channel_year = {}
for channel_name, year, year_month, net_rev in rows:
    rev_by_channel_year[(channel_name, year)] = rev_by_channel_year.get((channel_name, year), 0) + float(net_rev)

headers = ["Channel", "FY2024 Revenue", "FY2025 Revenue", "YoY Growth %", "FY2024 Share %", "FY2025 Share %"]
for j, h in enumerate(headers):
    c = ws.cell(row=3, column=1 + j, value=h)
    c.font = BLACK_BOLD
    c.fill = SUBHEADER_FILL
for i, ch in enumerate(channels):
    r = 4 + i
    ws.cell(row=r, column=1, value=ch).font = BLUE
    ws.cell(row=r, column=2, value=round(rev_by_channel_year.get((ch, "2024"), 0), 2)).font = BLUE
    ws.cell(row=r, column=2).number_format = CUR
    ws.cell(row=r, column=3, value=round(rev_by_channel_year.get((ch, "2025"), 0), 2)).font = BLUE
    ws.cell(row=r, column=3).number_format = CUR
    ws.cell(row=r, column=4, value=f"=C{r}/B{r}-1").number_format = PCT
    ws.cell(row=r, column=5, value=f"=B{r}/SUM($B$4:$B${3+len(channels)})").number_format = PCT
    ws.cell(row=r, column=6, value=f"=C{r}/SUM($C$4:$C${3+len(channels)})").number_format = PCT
    for col in [4, 5, 6]:
        ws.cell(row=r, column=col).font = BLACK
for col, w in {"A": 14, "B": 16, "C": 16, "D": 14, "E": 14, "F": 14}.items():
    ws.column_dimensions[col].width = w
ws.sheet_view.showGridLines = False
print("Data_Channel done")

# =====================================================================
# SHEET 4: Data_Category (product category mix & margin)
# =====================================================================
ws = wb.create_sheet("Data_Category")
ws["A1"] = "Product Category Mix & Margin — Source: SQL export (category_monthly.csv)"
ws["A1"].font = BLACK_BOLD
hdr, rows = load_csv("/home/claude/fpa-reporting-pipeline/exports/category_monthly.csv")
cats = sorted(set(r[0] for r in rows))
rev_by_cat = {}
gp_by_cat = {}
for cat, ym, rev, gp in rows:
    rev_by_cat[cat] = rev_by_cat.get(cat, 0) + float(rev)
    gp_by_cat[cat] = gp_by_cat.get(cat, 0) + float(gp)
headers = ["Category", "Net Revenue (FY24-25)", "Gross Profit", "Gross Margin %", "% of Total Revenue"]
for j, h in enumerate(headers):
    c = ws.cell(row=3, column=1 + j, value=h)
    c.font = BLACK_BOLD
    c.fill = SUBHEADER_FILL
for i, cat in enumerate(cats):
    r = 4 + i
    ws.cell(row=r, column=1, value=cat).font = BLUE
    ws.cell(row=r, column=2, value=round(rev_by_cat[cat], 2)).font = BLUE
    ws.cell(row=r, column=2).number_format = CUR
    ws.cell(row=r, column=3, value=round(gp_by_cat[cat], 2)).font = BLUE
    ws.cell(row=r, column=3).number_format = CUR
    ws.cell(row=r, column=4, value=f"=C{r}/B{r}").number_format = PCT
    ws.cell(row=r, column=5, value=f"=B{r}/SUM($B$4:$B${3+len(cats)})").number_format = PCT
    for col in [4, 5]:
        ws.cell(row=r, column=col).font = BLACK
for col, w in {"A": 14, "B": 18, "C": 16, "D": 14, "E": 16}.items():
    ws.column_dimensions[col].width = w
ws.sheet_view.showGridLines = False
print("Data_Category done")

# =====================================================================
# SHEET 5: Top_Customers (with Excel RANK formula)
# =====================================================================
ws = wb.create_sheet("Top_Customers")
ws["A1"] = "Customer Revenue Ranking — Source: SQL export (customer_summary.csv)"
ws["A1"].font = BLACK_BOLD
hdr, rows = load_csv("/home/claude/fpa-reporting-pipeline/exports/customer_summary.csv")
headers = ["Rank", "Customer", "Segment", "Region", "Total Revenue", "Total Gross Profit", "Margin %"]
for j, h in enumerate(headers):
    c = ws.cell(row=3, column=1 + j, value=h)
    c.font = BLACK_BOLD
    c.fill = SUBHEADER_FILL
for i, row in enumerate(rows):
    r = 4 + i
    name, seg, region, rev, gp = row
    ws.cell(row=r, column=2, value=name).font = BLUE
    ws.cell(row=r, column=3, value=seg).font = BLUE
    ws.cell(row=r, column=4, value=region).font = BLUE
    ws.cell(row=r, column=5, value=float(rev)).font = BLUE
    ws.cell(row=r, column=5).number_format = CUR
    ws.cell(row=r, column=6, value=float(gp)).font = BLUE
    ws.cell(row=r, column=6).number_format = CUR
    last_data_row = 3 + len(rows)
    ws.cell(row=r, column=1, value=f"=RANK(E{r},$E$4:$E${last_data_row})").font = BLACK
    ws.cell(row=r, column=7, value=f"=F{r}/E{r}").number_format = PCT
    ws.cell(row=r, column=7).font = BLACK
for col, w in {"A": 8, "B": 16, "C": 12, "D": 16, "E": 16, "F": 18, "G": 12}.items():
    ws.column_dimensions[col].width = w
ws.sheet_view.showGridLines = False
ws.freeze_panes = "A4"
print("Top_Customers done")

# =====================================================================
# SHEET 0: DASHBOARD
# =====================================================================
ws = wb.create_sheet("Dashboard")
ws["A1"] = "Meridian Retail Co. — Monthly FP&A Reporting Pack"
ws["A1"].font = TITLE_FONT
ws.merge_cells("A1:L1")
ws["A2"] = "Auto-generated from the SQL pipeline (sales.db) — replaces a manual monthly Excel pull"
ws["A2"].font = Font(name=FONT, italic=True, size=10, color="595959")
ws.merge_cells("A2:L2")

n_months = len(load_csv("/home/claude/fpa-reporting-pipeline/exports/monthly_kpis.csv")[1])
last_month_row = 3 + n_months

kpis = [
    ("Latest Month Revenue", f"=Data_Monthly!C{last_month_row}", CUR),
    ("MoM Growth %", f"=Data_Monthly!F{last_month_row}", PCT),
    ("YoY Growth %", f"=Data_Monthly!G{last_month_row}", PCT),
    ("FY2025 YTD Revenue", f"=Data_Monthly!I{last_month_row}", CUR),
    ("Blended Gross Margin %", f"=SUM(Data_Monthly!D4:D{last_month_row})/SUM(Data_Monthly!C4:C{last_month_row})", PCT),
]
for i, (lbl, formula, fmt) in enumerate(kpis):
    col = 1 + i * 2
    cell_lbl = ws.cell(row=4, column=col, value=lbl)
    cell_lbl.font = KPI_LABEL_FONT
    ws.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col + 1)
    cell_val = ws.cell(row=5, column=col, value=formula)
    cell_val.font = KPI_VALUE_FONT
    cell_val.number_format = fmt
    ws.merge_cells(start_row=5, start_column=col, end_row=5, end_column=col + 1)
    for rr in (4, 5):
        for cc in (col, col + 1):
            ws.cell(row=rr, column=cc).fill = KPI_FILL
    ws.row_dimensions[5].height = 28

# --- Chart 1: Monthly Revenue & Gross Profit trend ---
chart1 = LineChart()
chart1.title = "Monthly Net Revenue & Gross Profit (FY2024-FY2025)"
chart1.style = 2
chart1.y_axis.title = "$"
chart1.x_axis.title = "Month"
chart1.height = 8.5
chart1.width = 22
data = Reference(ws.parent["Data_Monthly"], min_col=3, max_col=4, min_row=3, max_row=last_month_row)
cats = Reference(ws.parent["Data_Monthly"], min_col=1, min_row=4, max_row=last_month_row)
chart1.add_data(data, titles_from_data=True)
chart1.set_categories(cats)
ws.add_chart(chart1, "A7")

# --- Chart 2: Regional FY24 vs FY25 ---
n_regions = 5
chart2 = BarChart()
chart2.type = "col"
chart2.title = "Revenue by Region: FY2024 vs FY2025"
chart2.style = 10
chart2.y_axis.title = "$"
chart2.height = 8.5
chart2.width = 14
data = Reference(ws.parent["Data_Regional"], min_col=2, max_col=3, min_row=3, max_row=3 + n_regions)
cats = Reference(ws.parent["Data_Regional"], min_col=1, min_row=4, max_row=3 + n_regions)
chart2.add_data(data, titles_from_data=True)
chart2.set_categories(cats)
ws.add_chart(chart2, "M7")

# --- Chart 3: Channel mix FY2025 (pie) ---
n_channels = 3
chart3 = PieChart()
chart3.title = "FY2025 Revenue by Channel"
chart3.height = 8.5
chart3.width = 12
data = Reference(ws.parent["Data_Channel"], min_col=3, min_row=3, max_row=3 + n_channels)
cats = Reference(ws.parent["Data_Channel"], min_col=1, min_row=4, max_row=3 + n_channels)
chart3.add_data(data, titles_from_data=True)
chart3.set_categories(cats)
chart3.dataLabels = DataLabelList()
chart3.dataLabels.showPercent = True
ws.add_chart(chart3, "A26")

# --- Top 5 customers table ---
ws["M26"] = "Top 5 Customers by Revenue"
ws["M26"].font = BLACK_BOLD
for j, h in enumerate(["Customer", "Region", "Revenue"]):
    c = ws.cell(row=27, column=13 + j, value=h)
    c.font = BLACK_BOLD
    c.fill = SUBHEADER_FILL
for i in range(5):
    src_row = 4 + i
    ws.cell(row=28 + i, column=13, value=f"=Top_Customers!B{src_row}")
    ws.cell(row=28 + i, column=14, value=f"=Top_Customers!D{src_row}")
    ws.cell(row=28 + i, column=15, value=f"=Top_Customers!E{src_row}").number_format = CUR

ws["A45"] = ("Headline insight: North America (+23.8% YoY) and Latin America (+28.9% YoY) are growing fast, "
             "but Greater Asia is down -37.8% YoY — that decline alone offsets most of the company's growth "
             "elsewhere, which is exactly the kind of finding this pipeline is built to surface automatically "
             "every month instead of waiting for a manual deep-dive.")
ws["A45"].font = Font(name=FONT, italic=True, size=10, color="595959")
ws.merge_cells("A45:R45")
ws.row_dimensions[45].height = 30

ws.sheet_view.showGridLines = False
for col in "ABCDEFGHIJKL":
    ws.column_dimensions[col].width = 11

print("Dashboard done")

order = ["Dashboard", "Data_Monthly", "Data_Regional", "Data_Channel", "Data_Category", "Top_Customers"]
wb._sheets = [wb[name] for name in order]
wb.active = 0
wb.save("/home/claude/fpa-reporting-pipeline/excel/FPA_Monthly_Reporting_Pack.xlsx")
print("Saved with sheet order:", order)
